import openpyxl
import json

# Open the JSON file
with open('config.json', 'r') as f:
    # Load the JSON data into a Python object
    config = json.load(f)

def apply_layout():
    # Load Excel workbook
    wb = openpyxl.load_workbook("timeline.xlsx")
    sheet = wb.active

    # Iterate through the first 2 columns in the sheet
    for cell in sheet.iter_cols(min_col=1, max_col=2):
        for i in range(len(cell)):
            # Make column size fit text length
            if len(cell[i].value) + 1 > sheet.column_dimensions[chr(ord('@') + cell[i].column)].width:
                sheet.column_dimensions[chr(ord('@') + cell[i].column)].width = len(cell[i].value) + 1

    wb.save("timeline.xlsx")
    
    # Order the issues with selected config
    if config['layout']['order'] == "alphabetical":
        sort_rows_alphabetical('A')

    set_colors()

def sort_rows_alphabetical(column_letter):
    # Open the workbook
    wb = openpyxl.load_workbook("timeline.xlsx")
    sheet = wb.active

    # Get the rows as a list of tuples (each tuple represents a row)
    rows = list(sheet.rows)
    rows.pop(0)

    # # Get the index of the column you want to sort by
    column_index = ord(column_letter.lower()) - ord('a')

    # # Sort the rows alphabetically by the text in the specified column
    rows.sort(key=lambda row: row[column_index].value.lower())

    # Clear the sheet
    for i in range(2, sheet.max_row + 1):
        sheet.delete_rows(2)

    # Write the sorted rows back to the sheet
    for row in rows:
        sheet.append([cell.value for cell in row])
    
    wb.save("timeline.xlsx")
    
def set_colors():
    # Open the workbook
    wb = openpyxl.load_workbook("timeline.xlsx")
    sheet = wb.active

    # Color all cells
    for cell in sheet.iter_cols(min_col=3):
        for i in range(len(cell)):
            if cell[i].value in config['authors']:
                sheet[chr(ord('@') + cell[i].column) + str(cell[i].row)].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=config['authors'][cell[i].value])
                sheet[chr(ord('@') + cell[i].column) + str(cell[i].row)].value = None
                
    wb.save("timeline.xlsx")